# telemetry-analysis/telemetry_analysis/mileage_spreadsheet.py
"""
Imports data from Mileage Spreadsheet and matches that to EXIF data from images
taken at each fuel stop.  Typically, one of the images at each fuel stop will
show the vehicle's odometer.

Matching of EXIF image data to spreadsheet data is based on both date/time and location info.

Output includes spreadsheet and image file info for each vehicle.

See telemetry-analysis/docs/mileage_spreadsheet.md for more info.
"""
if __name__ == "__main__":
    import nbimporter

import json
from pathlib import Path
from argparse import ArgumentParser
from itertools import count
from dateutil import parser
from datetime import datetime, timedelta
from pytz import timezone
from openpyxl import load_workbook
from rich.pretty import pprint

from .pictures import image_directory_to_exif

# DEFAULT Python File Name - mileage_{vin}
DEFAULT_PYTHON_DIRECTORY = Path.home() / Path("Dropbox/src/telemetry-analysis/private")
DEFAULT_XLSX = Path.home() / Path("Dropbox/RV/Mileage.xlsx")
DEFAULT_IMAGE_DIRECTORY = Path.home() / Path("telemetry-data/fuel-images")
DEFAULT_DATA_DIRECTORY = Path.home() / Path("telemetry-data")
DEFAULT_HOME_TIMEZONE = timezone('US/Central')
# 4 hours 
DEFAULT_MAXIMUM_TIME_DIFFERENCE = timedelta(seconds=(4*60*60))
ONE_DAY = timedelta(days=1.0)

def naive_datetime_to_aware_datetime(naive)->datetime:
    """
    Create a default home timezone aware datetime object
    """
    return DEFAULT_HOME_TIMEZONE.localize(naive)

def string_to_list(sheets:str) -> list:
    """
    take a string with comma separated sheet names and return
    a list of sheet names
    """
    return sheets if isinstance(sheets, list) else sheets.split(',')

def spreadsheet(spreadsheet:Path, sheet:str, vin:str, verbose=False) -> list:
    """
    Load a sheet from an Excel Spreadsheet and turn it into a list of
    mileage information entries
    """
    xlsx = load_workbook(filename=spreadsheet, read_only=True, data_only=True)

    if sheet not in xlsx.sheetnames:
        raise ValueError(f"Bad Spreadsheet {spreadsheet}: sheet {sheet} not in {xlsx.sheetnames}")

    ws = xlsx[sheet]

    rows = []
    for row in count(2):
        if verbose:
            print(".", end='')

        if not ws.cell(row=row, column=1).value:
            # empty cell means we passed the last valid row
            xlsx.close()
            if verbose:
                print(f"\n{spreadsheet}\nsheet '{sheet}': {row} rows read")
            return rows

        tmp_row = {
            'vin': vin,
            'Date': ws.cell(row=row, column=1).value,
            'MPG':  (ws.cell(row=row, column=2)).value,
            'Odometer': ws.cell(row=row, column=3).value,
            'Gallons': (ws.cell(row=row, column=4)).value,
            'PricePerGallon': ws.cell(row=row, column=5).value,
            'FuelBrand': ws.cell(row=row, column=6).value,
            'StationAddress': ws.cell(row=row, column=7).value,
        }

        if isinstance(tmp_row['Date'], datetime):
            # convert datetime object to a default home timezone aware datetime.
            tmp_row['Date'] = naive_datetime_to_aware_datetime(tmp_row['Date'])
        else:
            if verbose:
                print(f"Date ({tmp_row['Date']}) not datetime at row {row}")
            tmp_row['Date'] = None

        for column in ['MPG', 'Odometer', 'Gallons', 'PricePerGallon', ]:
            if not isinstance(tmp_row[column], float):
                try:
                    tmp_row[column] = float(tmp_row[column])
                except Exception as e:
                    if verbose:
                        print(f"\n{e}")
                        print(f"{column} ({tmp_row[column]}) not float at row {row}\n")
                    tmp_row['column'] = None

        rows.append(tmp_row)

    xlsx.close()

    return rows

def mileage_spreadsheet(
    sheets=None,
    vins=None,
    xlsx=DEFAULT_XLSX,
    image_directory=DEFAULT_IMAGE_DIRECTORY,
    python_directory=DEFAULT_PYTHON_DIRECTORY,
    verbose=False
)->dict:
    if len(vins) != len(sheets):
        raise ValueError(f"The number of vins ({len(vins)}) not equal to number of sheets ({len(sheets)})")

    mileage_spreadsheet_records = []
    for vin_sheet_index in range(len(sheets)):
        vin = vins[vin_sheet_index]
        sheet = sheets[vin_sheet_index]

        mileage_spreadsheet_records += spreadsheet(xlsx, sheet, vin, verbose=verbose)

    # convert to dictionary where:
    # key = (vin, record['Date'], record_number)
    #       where record_number = 0 is the first record in the list
    # value is record
    fuel_fill_data = {
        (record['vin'], record['Date'], record_number):  record
        for record_number, record in enumerate(mileage_spreadsheet_records) if record['Date']
    }

    if verbose:
        pprint(fuel_fill_data)

    return fuel_fill_data

def gps_file_information(file_name, verbose=False)->tuple:
    """
    file_name for a GPS Logger data file
    returns
    - tuple of (iso_ts_pre, iso_ts_post, first_location, last_location)
    where
    - iso_ts_pre is the iso_ts_pre timestamp from the first valid location record
    - iso_ts_post is the iso_ts_post from the last valid location record
    - first_location is the location from the first valid location record
    - last_location is the location from the last valid location record
    """
    if verbose:
        print(f"GPS input file {file_name}")

    iso_ts_pre = None
    iso_ts_post = None
    first_location = None
    last_location = None

    with open(file_name, "r") as json_input:
        for line_number, json_record in enumerate(json_input, start=1):
            try:
                record = json.loads(json_record)
            except json.decoder.JSONDecodeError as e:
                # improperly closed JSON file
                if verbose:
                    print(f"Corrupted GPS JSON info at line {line_number}: {e}")
                return iso_ts_pre, iso_ts_post, first_location, last_location
            # {
            #   "command_name": "NMEA_GNGNS",
            #   "obd_response_value": {
            #       "time": "18:14:57",
            #       "lat": "29.51983417",
            #       "NS": "N",
            #       "lon": "-98.573058",
            #       "EW": "W",
            #       "posMode": "AA     ",
            #       "numSV": "13",
            #       "HDOP": "1.96",
            #       "alt": "292.4",
            #       "sep": "-22.8",
            #       "diffAge": null,
            #       "diffStation": null
            #   },
            #   "iso_ts_pre": "2023-03-24T22:39:33.542856+00:00",
            #   "iso_ts_post": "2023-03-24T22:39:33.554321+00:00"
            # }
            if record['command_name'] == 'NMEA_GNGNS' and record['obd_response_value']['lat']:
                if not iso_ts_pre:
                    iso_ts_pre = parser.isoparse(record['iso_ts_pre'])
                    first_location = {
                        'time': record['obd_response_value']['time'],
                        'lat': record['obd_response_value']['lat'],
                        'NS': record['obd_response_value']['NS'],
                        'lon': record['obd_response_value']['lon'],
                        'EW': record['obd_response_value']['EW'],
                        'alt': record['obd_response_value']['alt'],
                    }
                iso_ts_post = parser.isoparse(record['iso_ts_post'])
                last_location = {
                    'time': record['obd_response_value']['time'],
                    'lat': record['obd_response_value']['lat'],
                    'NS': record['obd_response_value']['NS'],
                    'lon': record['obd_response_value']['lon'],
                    'EW': record['obd_response_value']['EW'],
                    'alt': record['obd_response_value']['alt'],
                }


    return iso_ts_pre, iso_ts_post, first_location, last_location

def gps_logger_data(data_directory=DEFAULT_DATA_DIRECTORY, verbose=False)->dict:
    """
    For each GPS Logger data file, place the file name with the first and last location
    into a dictionary.

    dictionary key is the tuple (first_location_record_iso_ts_pre, last_location_record_iso_ts_post)
    dictionary data is the dictionary of file_name, first_location_record, last_location_record
    """
    root = Path(data_directory)
    gps_files = [ str(gps_file) for gps_file in root.rglob("NMEA-[0-9]*-utc.json") if gps_file.is_file()]
    gps_files += [ str(gps_file) for gps_file in root.rglob("*[0-9]-gps-[0-9]*.json") if gps_file.is_file()]

    gps_data = {}

    for gps_file in gps_files:
        iso_ts_pre, iso_ts_post, first_location, last_location = gps_file_information(gps_file, verbose=verbose)

        if iso_ts_pre:
            gps_data[(iso_ts_pre, iso_ts_post)] = {
                'iso_ts_pre': iso_ts_pre,
                'iso_ts_post': iso_ts_post,
                'file_name': gps_file,
                'first_location': first_location,
                'last_location': last_location,
            }

    if verbose:
        print("gps_logger_data() return data")
        pprint(gps_data)

    return gps_data

def strip_units_from_value(obd_response_value: str):
    """
    OBD Logger response values often look like string "58373.3 mile"
    Transform the value into a float 58373.3
    """
    if isinstance(obd_response_value, str):
        return float((obd_response_value.split(' '))[0])
    elif isinstance(obd_response_value, int):
        return float(obd_response_value)
    
    return None

def obd_file_information(file_name:Path, verbose=False) -> tuple:
    """
    file_name for a OBD Logger data file
    returns
    - tuple of (iso_ts_pre, iso_ts_post, vin)
    where
    - iso_ts_pre is the iso_ts_pre timestamp from the first valid ODOMETER/FUEL_LEVEL record
    - iso_ts_post is the iso_ts_post from the last valid ODOMETER/FUEL_LEVEL record
    - first_FUEL_LEVEL is the FUEL_LEVEL from the first valid FUEL_LEVEL record
    - last_FUEL_LEVEL is the FUEL_LEVEL from the last valid FUEL_LEVEL record
    - first_ODOMETER is the ODOMETER from the first valid ODOMETER record
    - last_ODOMETER is the ODOMETER from the last valid ODOMETER record
    """
    if verbose:
        print(f"Reading OBD Logger file: {file_name}")

    iso_ts_pre = None
    iso_ts_post = None
    first_ODOMETER = None
    last_ODOMETER = None
    first_FUEL_LEVEL = None
    last_FUEL_LEVEL = None

    with open(file_name, "r") as json_input:
        for line_number, json_record in enumerate(json_input, start=1):
            try:
                record = json.loads(json_record)
            except json.decoder.JSONDecodeError as e:
                # improperly closed JSON file
                if verbose:
                    print(f"Corrupted OBD JSON info at line {line_number}: {e}")
                return iso_ts_pre, iso_ts_post, first_ODOMETER, last_ODOMETER, first_FUEL_LEVEL, last_FUEL_LEVEL

            if record['command_name'] in ['ODOMETER', 'FUEL_LEVEL', ]:
                if not iso_ts_pre:
                    iso_ts_pre = parser.isoparse(record['iso_ts_pre'])

                iso_ts_post = parser.isoparse(record['iso_ts_post'])

            if record['command_name'] == 'ODOMETER' and record['obd_response_value']:
                if not first_ODOMETER:
                    first_ODOMETER = strip_units_from_value(record['obd_response_value'])
                if temporary_last_ODOMETER := strip_units_from_value(
                    record['obd_response_value']
                ):
                    last_ODOMETER = temporary_last_ODOMETER

            if record['command_name'] == 'FUEL_LEVEL' and record['obd_response_value']:
                if not first_FUEL_LEVEL:
                    first_FUEL_LEVEL = strip_units_from_value(record['obd_response_value'])
                if temporary_last_FUEL_LEVEL := strip_units_from_value(
                    record['obd_response_value']
                ):
                    last_FUEL_LEVEL = temporary_last_FUEL_LEVEL

    return iso_ts_pre, iso_ts_post, first_ODOMETER, last_ODOMETER, first_FUEL_LEVEL, last_FUEL_LEVEL


def obd_logger_data(vins:list, data_directory=DEFAULT_DATA_DIRECTORY, verbose=False)->dict:
    """
    For each OBD Logger data file, place the file name with the first and last iso_ts_pre/iso_ts_post
    ODOMETER and FUEL_LEVEL into a dictionary.

    dictionary key is the tuple (ODOMETER/FUEL_LEVEL_record_iso_ts_pre, ODOMETER/FUEL_LEVEL_record_iso_ts_post)
    dictionary data is the dictionary of file_name, first_ODOMETER_record, last_ODOMETER_record,
    first_FUEL_LEVEL_record, last_FUEL_LEVEL_record
    """
    root = Path(data_directory)

    obd_data = {}

    for vin in vins:
        obd_files = [str(gps_file) for gps_file in root.rglob(f"*{vin}*.json") if gps_file.is_file() if 'integrated' not in str(gps_file)]


        for obd_file in obd_files:
            iso_ts_pre, iso_ts_post, first_ODOMETER, last_ODOMETER, first_FUEL_LEVEL, last_FUEL_LEVEL = obd_file_information(
                obd_file, verbose=verbose
            )
            if iso_ts_pre:
                obd_data[(vin, iso_ts_pre, iso_ts_post)] = {
                    'vin': vin,
                    'iso_ts_pre': iso_ts_pre,
                    'iso_ts_post': iso_ts_post,
                    'file_name': obd_file,
                    'first_ODOMETER': first_ODOMETER,
                    'last_ODOMETER': last_ODOMETER,
                    'first_FUEL_LEVEL': first_FUEL_LEVEL,
                    'last_FUEL_LEVEL': last_FUEL_LEVEL,
                }

    if verbose:
        print("obd_logger_data() return data")
        pprint(obd_data)

    return obd_data

def fuel_image_data(image_directory=DEFAULT_IMAGE_DIRECTORY, verbose=False)->dict:
    # sourcery skip: use-named-expression, use-next
    """
    Gets image data (timestamp and location) from images (*.jgp) in image_directory.
    Returns dictionary with:
        key: (file_timestamp, exif_time)
    """
    image_data = image_directory_to_exif(image_directory=image_directory, verbose=verbose)

    return_value = {}

    for k, v in image_data.items():
        first_key_part = None
        for dt in ['aware_gps_datetime', 'DateTime', 'DateTimeOriginal', 'DateTimeDigitized', ]:
            # find the first valid datetime object and use it in the key
            # where the key is (datetime.datetime, file_name)
            if dt in image_data[k] and isinstance(image_data[k][dt], datetime):
                first_key_part = dt
                break

        if first_key_part:
            return_value[(v[first_key_part], k)] = v
        else:
            if verbose:
                print(f"No valid timestamp available in image exif: {image_data[k]}")
            return_value[(datetime(1,1,1), k)] = v

    return return_value

def combine_data(
        vins, fuel_fill_data, location_data, engine_data, fuel_fill_picture_data,
        maximum_time_difference=DEFAULT_MAXIMUM_TIME_DIFFERENCE,
        verbose=False
) -> dict:
    """
    Integrate data so that all available data can be connected fuel fill events found in the spreadsheet data.
    """
    combined_data = {}
    for vin in vins:
        if verbose:
            print(f"combine_data({vin})")

        # limit fuel fill data to this vin
        fuel = {k: v for k, v in fuel_fill_data.items() if k[0] == vin}

        # limit engine_data to this vin
        engine = {k: v for k, v in engine_data.items() if k[0] == vin}

        matching_engine = {}
        for filling in fuel:
            for k, v in engine.items():
                # first_iso_ts_pre must be same day
                if filling[1] < k[1] and abs(k[1] - filling[1]) <= ONE_DAY:
                    matching_engine[k] = v
                    if verbose:
                        print(f"fuel_fill_data[{filling}] same day as first_iso_ts_pre engine_data[{k}]")
                # or last_iso_ts_post must be same day
                elif filling[1] < k[2] and abs(k[2] - filling[1]) <= ONE_DAY:
                    matching_engine[k] = v
                    if verbose:
                        print(f"fuel_fill_data[{filling}] same day as last_iso_ts_post engine_data[{k}]")



    return combined_data

def gather_data(
        sheets=None,
        vins=None,
        xlsx=DEFAULT_XLSX,
        image_directory=DEFAULT_IMAGE_DIRECTORY,
        python_directory=DEFAULT_PYTHON_DIRECTORY,
        data_directory=DEFAULT_DATA_DIRECTORY,
        verbose=False
):
    """
    Gather data from images, spreadsheets, telemetry_obd.obd_logger
    and telemetry_gps.gps_logger to create mapping between the four.

    The mapping between the four sources of data will be placed into
    a python program file where the dictionary containing the mapping
    can be imported by other python programs.
    """
    # mileage_spreadsheet() sample output:
        # ('3FTTW8F97PRA00000', datetime.datetime(2023, 8, 23, 0, 0, tzinfo=<DstTzInfo 'US/Central' CDT-1 day, 19:00:00 DST>), 0): {
        #     'vin': '3FTTW8F97PRA00000',
        #     'Date': datetime.datetime(2023, 8, 23, 0, 0, tzinfo=<DstTzInfo 'US/Central' CDT-1 day, 19:00:00 DST>),
        #     'MPG': None,
        #     'Odometer': 76.1,
        #     'Gallons': 14.5,
        #     'PricePerGallon': None,
        #     'FuelBrand': None,
        #     'StationAddress': 'Jordan Ford',
        #     'column': None
        # },
    fuel_fill_data = mileage_spreadsheet(
        sheets=sheets,
        vins=vins,
        xlsx=Path(xlsx),
        image_directory=Path(image_directory),
        python_directory=Path(python_directory),
        verbose=verbose
    )
    fuel_fill_data = dict(sorted(fuel_fill_data.items()))

    # gps_logger_data() sample output:
        # (datetime.datetime(2000, 1, 1, 0, 0, 55, 892776, tzinfo=tzutc()), datetime.datetime(2000, 1, 1, 0, 7, 41, 897955, tzinfo=tzutc())): {
        #     'iso_ts_pre': datetime.datetime(2000, 1, 1, 0, 0, 55, 892776, tzinfo=tzutc()),
        #     'iso_ts_post': datetime.datetime(2000, 1, 1, 0, 7, 41, 897955, tzinfo=tzutc()),
        #     'file_name': 'C:\\Users\\runar\\telemetry-data\\telemetry-gps\\data\\NMEA-20000101000034-utc.json',
        #     'first_location': {
        #         'time': '16:59:56',
        #         'lat': '29.51895683',
        #         'NS': 'N',
        #         'lon': '-98.4627545',
        #         'EW': 'W',
        #         'alt': '242.6'
        #     },
        #     'last_location': {
        #         'time': '17:06:41',
        #         'lat': '29.51488633',
        #         'NS': 'N',
        #         'lon': '-98.464322',
        #         'EW': 'W',
        #         'alt': '242.0'
        #     }
        # },
    location_data = gps_logger_data(
        data_directory=data_directory,
        verbose=verbose
    )
    location_data = dict(sorted(location_data.items()))

    # obd_logger_data() sample output:
        # ('3FTTW8F97PRA00000', datetime.datetime(2023, 9, 1, 17, 11, 30, 545411, tzinfo=tzutc()), datetime.datetime(2023, 9, 1, 18, 14, 26, 839080, tzinfo=tzutc())): {
        #     'vin': '3FTTW8F97PRA00000',
        #     'iso_ts_pre': datetime.datetime(2023, 9, 1, 17, 11, 30, 545411, tzinfo=tzutc()),
        #     'iso_ts_post': datetime.datetime(2023, 9, 1, 18, 14, 26, 839080, tzinfo=tzutc()),
        #     'file_name': 'C:\\Users\\runar\\telemetry-data\\data\\3FTTW8F97PRA00000\\3FTTW8F97PRA00000-0000000001.json',
        #     'first_ODOMETER': 3347.2,
        #     'last_ODOMETER': 4118.1,
        #     'first_FUEL_LEVEL': 76.07843137254902,
        #     'last_FUEL_LEVEL': 51.372549019607845
        # },
    engine_data = obd_logger_data(
        vins=vins,
        data_directory=data_directory,
        verbose=verbose
    )
    engine_data = dict(sorted(engine_data.items()))

    # fuel_image_data() sample output:
        # (datetime.datetime(2017, 6, 11, 15, 43, 26, tzinfo=<UTC>), 'C:\\Users\\runar\\telemetry-data\\fuel-images\\20170611-2017-06-11 10.43.26.jpg'): {
        #     'GPSInfo': {
        #         'time': '15:43:26',
        #         'date': '2017/06/11',
        #         'aware_gps_datetime': datetime.datetime(2017, 6, 11, 15, 43, 26, tzinfo=<UTC>),
        #         'lat': 29.6512,
        #         'ns': 'N',
        #         'lon': 97.59375,
        #         'ew': 'W',
        #         'alt': 93.71814671814671
        #     },
    fuel_fill_picture_data = fuel_image_data(image_directory=image_directory, verbose=verbose)
    fuel_fill_picture_data = dict(sorted(fuel_fill_picture_data.items()))


def command_line_options()->dict:
    parser = ArgumentParser(prog="mileage_spreadsheet", description="Transform mileage spreadsheet and image EXIF data into importable python file.")

    parser.add_argument(
        "--sheets",
        help="""
            An ordered list of spreadsheet sheet names separated by commas
            to include in python file generation.
            e.g. "JeepWranglerRubicon,F450,EcoSport,Maverick"
        """,
    )

    parser.add_argument(
        "--vins",
        help="""
            Vehicle VIN corresponding to selected sheet in spreadsheet.
            e.g. 'C4HJWCG5DL0000,FT8W4DT5HED00000,MAJ6S3KL0KC2000000,3FTTW8F97PRA00000'.
        """,
    )

    parser.add_argument(
        "--xlsx",
        help=f"""
            Excel mileage spreadsheet file name/path.
            Defaults to "{DEFAULT_XLSX}".
        """,
        default=DEFAULT_XLSX,
    )

    parser.add_argument(
        "--image_directory",
        help=f"""
        Directory where the images of fuel stops are stored.
        Defaults to "{DEFAULT_IMAGE_DIRECTORY}"
        """,
        default=DEFAULT_IMAGE_DIRECTORY,
    )

    parser.add_argument(
        "--data_directory",
        help=f"""
        Directory where the images of fuel stops are stored.
        Defaults to "{DEFAULT_DATA_DIRECTORY}"
        """,
        default=DEFAULT_DATA_DIRECTORY,
    )

    parser.add_argument(
        "--python_directory",
        help=f"""Python output file directory.
                Can be either a full or relative path name.
                Defaults to '{DEFAULT_PYTHON_DIRECTORY}'.
                """,
        default=DEFAULT_PYTHON_DIRECTORY,
    )

    parser.add_argument(
        "--verbose",
        help="Turn verbose output on. Default is off.",
        default=False,
        action='store_true'
    )

    return vars(parser.parse_args())

if __name__ == "__main__":
    args = command_line_options()

    if args['verbose']:
        print(f"sheets: {args['sheets']}")
        print(f"vins: {args['vins']}")
        print(f"xlsx: {args['xlsx']}")
        print(f"image_directory: {args['image_directory']}")
        print(f"python_directory: {args['python_directory']}")
        print(f"data_directory: {args['data_directory']}")
        print(f"verbose: {args['verbose']}")

    sheets = args['sheets']
    if isinstance(sheets, str):
        sheets = string_to_list(sheets)

    vins = args['vins']
    if isinstance(vins, str):
        vins = string_to_list(vins)

    xlsx = args['xlsx']
    image_directory = args['image_directory']
    python_directory = args['python_directory']
    data_directory = args['data_directory']
    verbose = args['verbose']

    gather_data(
        sheets=sheets,
        vins=vins,
        xlsx=xlsx,
        image_directory=image_directory,
        python_directory=python_directory,
        data_directory=data_directory,
        verbose=verbose
    )
